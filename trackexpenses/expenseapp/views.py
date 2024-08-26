from django.shortcuts import render,redirect
from django.contrib.auth.decorators import login_required
from .models import Category,Expense
from income.models import Income,Source
from django.contrib import messages
from django.core.paginator import Paginator
import json #you  can get everything that the user is trying to search for
from django.http import JsonResponse,HttpResponse
#we get data as json and convert into python dictionarty in the same way to send  data back to user in json we use http response
# Create your views here.
import datetime
from django.core import serializers
from django.db.models import Sum
import openpyxl
from openpyxl.styles import Alignment

def search_expenses(request):
 if request.method=='POST':
  search_str=json.loads(request.body).get('searchText') #json.loads convert the json obtained from request.body to python dictionary
  expenses=Expense.objects.filter(amount__istartswith=search_str,owner=request.user) | Expense.objects.filter(description__icontains=search_str,owner=request.user) | Expense.objects.filter(date__istartswith=search_str,owner=request.user) | Expense.objects.filter(category__icontains=search_str,owner=request.user)
  # second parameter is used so that onnly the expense that particualar user has added to expense and not other user expenses will be displayed
  # in expenses variable we now have a query set

  data=expenses.values()
  return JsonResponse(list(data),safe=False) #safe=false allows that allows jsonresponse class to be pass something that is not dictionary easily
@login_required(login_url='/authentication/login')
def index(request):
 categories=  Category.objects.all()
 expenses=Expense.objects.filter(owner=request.user) #all the expenses of that particular user
 paginator=Paginator(expenses,3) #pass two parameters i.e one is the object that we want to paginate and another is number of item to be displayed in a page
 page_number=request.GET.get('page' ) #get current page number 
 #GET.get
 page_obj=paginator.get_page(page_number) # which data to display in that page number  
 totalpage=page_obj.paginator.num_pages; #num_pages gives total number of pages
 context={
  'expenses':expenses,
  'page_obj':page_obj,
  'totalpageList': [n+1 for n in range(totalpage)],
  'totalpage':totalpage
 }
 return render(request,'index.html',context)



def add_expense(request):
 categories = [
    ("Health", "Health"),
    ("Insurance", "Insurance"),
    ("Education", "Education"),
    ("Fuel", "Fuel"),
    ("Groceries", "Groceries"),
    ("Electricity", "Electricity"),
    ("Internet","Internet"),
    ("Airlines","Airlines"),
    ("Others","Others")
  ] 
# categories=  Category.objects.all()
 context={
  'categories':categories,
  'values':request.POST
 }
 if request.method=='GET':
   return render(request,'add_expense.html',context)

 if request.method=='POST':
  amount=request.POST['amount']
  description=request.POST['description']  
  date=request.POST['dateOfExpense'] 
  category=request.POST['category']

  if not amount :
   messages.error(request,"Amount is required")
   return render(request,'add_expense.html',context)
  if not amount.isdigit():
   messages.error(request, 'Amount should be a number.')
   return render(request, 'add_expense.html',context)
  if not amount.isdecimal():
    messages.error(request, 'Amount should be a number. You can include a decimal point (.) if necessary.')
    return render(request, 'add_expense.html', context)
  
  if not description :
    messages.error(request,"Description is required")
    return render(request,'add_expense.html',context)
  
  if not date :
    messages.error(request,"All the fields are required")
    return render(request,'add_expense.html',context)
  
  if not category :
    messages.error(request,"All the fields are required")
    return render(request,'add_expense.html',context)

  
  Expense.objects.create(owner=request.user,amount=amount,date=date,category=category,description=description) #if owner is not included then it shows eror because owner is defined inside in model and we need to save data inside the owner  too
  expenses=Expense.objects.filter(owner=request.user)
  context={
  'expenses':expenses
  }
  messages.success(request,'Expense saved succesfully!!')
  return redirect('expenses')
  # return render(request,'index.html',context)


#EDIT

def edit_expense(request,id):
 expense=Expense.objects.get(pk=id)
 #categories= Category.objects.all()
 categories = [
    ("Health", "Health"),
    ("Insurance", "Insurance"),
    ("Education", "Education"),
    ("Fuel", "Fuel"),
    ("Groceries", "Groceries"),
    ("Electricity", "Electricity"),
    ("Internet","Internet"),
    ("Airlines","Airlines"),
    ("Others","Others")
  ] 
 context={
  'expense':expense,
  'values':expense,
  'categories':categories
  
 }

 if request.method=='GET':
  return render(request,'edit_expense.html',context)
 if request.method=='POST':
  amount=request.POST['amount']
  description=request.POST['description']  
  date=request.POST['dateOfExpense'] 
  category=request.POST['category']

  if not amount :
   messages.error(request,"Amount is required")
   return render(request,'edit_expense.html',context)
  

  if not description :
    messages.error(request,"Description is required")
    return render(request,'edit_expense.html',context)
  if not date :
    messages.error(request,"All the fields are required")
    return render(request,'edit_expense.html',context)
  if not category :
    messages.error(request,"All the fields are required")
    return render(request,'edit_expense.html',context)
 
  

  expense.owner=request.user
  expense.amount= amount
  expense.date= date
  expense.category=category
  expense.description=description
  expense.save()
  messages.success(request,'Expense updated successfully')

  return redirect(index)
 
def delete_expense(request,id):
  expense=Expense.objects.get(pk=id)  
  expense.delete()
  messages.success(request,'Expense removed')
  return redirect(index)

def expense_category_summary(request):
 today_date=datetime.date.today()
 three_months_ago=today_date-datetime.timedelta(days=30*3) 
 expense=Expense.objects.filter(owner=request.user,date__gte=three_months_ago,date__lte=today_date)
 finalrep={}
   # Define a helper function to extract the category from an expense object
 def get_category(expense):
  return expense.category
 
 # Create a list of unique categories from the expenses
 category_list=list(set(map(get_category,expense)))

 # Define a helper function to calculate the total amount spent for a given category
 def get_expense_category_amount(category):
  amount=0;
  filtered_by_category=expense.filter(category=category) 
  for item in filtered_by_category:
   amount+=item.amount
  return amount

#iterate through each expense and itertate through each category and find total amount for each category
 for x in expense:
  for y in category_list:
   finalrep[y]=get_expense_category_amount(y) #total sum for that particular category
 total_sum=sum(finalrep.values()) #summ all the category i.e sum of all the expenses
 combined={'expense_category_data':finalrep,
            'total_sum':total_sum
        
     }

 return JsonResponse(combined,safe=False)

def stats(request):
 return render(request,'stats.html')




def income_source_summary(request):
    today_date = datetime.date.today()
    three_months_ago = today_date - datetime.timedelta(days=30*3) 
    income = Income.objects.filter(owner=request.user, date__gte=three_months_ago, date__lte=today_date)

    finalrep = {}
    source_list = set(income.values_list('source', flat=True))

    def get_income_source_amount(source):
        filtered_by_source = income.filter(source=source)
        amount = sum(item.amount for item in filtered_by_source)
        return amount

    for source in source_list:
        finalrep[source] = get_income_source_amount(source)

    total_sum = sum(finalrep.values())

    combined = {
        'income_category_data': finalrep,
        'total_sum': total_sum
    }

    return JsonResponse(combined, safe=False)

def export_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Expenses_{datetime.date.today().strftime("%Y-%m-%d")}.xlsx'
    wb = openpyxl.Workbook()  # workbook as excelfile
    ws = wb.active
    ws.title = 'Expenses'
    row_num = 1
    columns = ['Amount', 'Description', 'Category', 'Date']
    ws.append(columns)

    expenses = Expense.objects.filter(owner=request.user).values_list('amount', 'description', 'category', 'date')

    for expense in expenses:
        ws.append(expense)
    column_widths = [15, 30, 20, 20]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = width

    # Format the 'Date' column as date
    date_col_idx = columns.index('Date') + 1
    for row in ws.iter_rows(min_row=2, min_col=date_col_idx, max_col=date_col_idx):
        for cell in row:
            cell.number_format = 'YYYY-MM-DD'
            cell.alignment = Alignment(horizontal='center')
    

    wb.save(response)
    return response